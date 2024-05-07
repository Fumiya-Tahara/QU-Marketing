import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side

def makeSetOfChoices(path, turnBackPoint):
  """
  path: 読み込むCSVファイルのパス
  turnBackPoint: 表の折り返し地点
  """
  file = pd.read_csv(path, encoding="shift-jis")
  headers = file.columns.tolist()

  records = []
  for index, row in file.iterrows():
    row_values = row.values.tolist()
    records.append(row_values)

  workbook = Workbook()
  sheet = workbook.active

  blockCount = 0
  rowStart = 1
  columnStart = 1
  border = Side(style='thin', color='000000')
  borderAro = Border(top=border, bottom=border, left=border, right=border)
  for i in range(len(records)//2):
    sheet.cell(row=rowStart, column=columnStart, value=i+1)
    emptyCell = sheet.cell(row=rowStart+1, column=columnStart)
    headerA = sheet.cell(row=rowStart+1, column=columnStart+1, value="A")
    headerB = sheet.cell(row=rowStart+1, column=columnStart+2, value="B")
    emptyCell.fill = PatternFill("solid", fgColor="D3D3D3")
    headerA.fill = PatternFill("solid", fgColor="D3D3D3")
    headerB.fill = PatternFill("solid", fgColor="D3D3D3")
    emptyCell.border = borderAro
    headerA.border = borderAro
    headerB.border = borderAro
    for j in range(len(headers)):
      rowNum = rowStart+j+2
      headerCell = sheet.cell(row=rowNum, column=columnStart, value=headers[j])
      aCell = sheet.cell(row=rowNum, column=columnStart+1, value=records[2*i][j])
      bCell = sheet.cell(row=rowNum, column=columnStart+2, value=records[2*i+1][j])
      headerCell.fill = PatternFill("solid", fgColor="D3D3D3")
      headerCell.border = borderAro
      aCell.border = borderAro
      bCell.border = borderAro
    columnStart += 4
    blockCount += 1
    if blockCount % turnBackPoint == 0:
      rowStart += len(headers) + 3
      columnStart = 1

  workbook.save('output.xlsx')
